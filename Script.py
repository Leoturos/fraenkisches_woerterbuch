import re
import json
import pandas as pd
from pathlib import Path
import csv
import time
import openpyxl
from openpyxl import Workbook
from tkinter import *
import tkinter.filedialog as FD
import tkinter

fenster= Tk()
fenster.title("Ableitungschecker")

global df
df=pd.DataFrame()
# Listen der Präfixe, Zirkumfixe und Suffixe
prefixes = [
    "a", "an", "ar", "ab", "aber", "after", "anti", "auf", "aus", "auto", "be", "bei", "bio", "de", "des", "dis",
    "durch", "ein", "emp", "ent", "entgegen", "er", "erz", "ex", "fehl", "fest", "fort", "ge", "gegen", "geo",
    "graf", "her", "herunter", "hin", "hinter", "hyper", "ident", "in", "im", "il", "ir", "inne", "inter", "ko", "kol",
    "kom", "kon", "kor", "kund", "los", "makro", "maxi", "mega", "mikro", "mini", "miss", "mit", "mono", "multi",
    "nach", "nano", "naut", "neo", "non", "para", "pflichtig", "phil", "phob", "poly", "post", "prä", "pro", "proto",
    "pseudo", "quasi", "re", "riesen", "rück", "schwieger", "semi", "stereo", "stief", "tele", "therm", "thermo",
    "trans", "ultra", "um", "un", "unter", "ur", "ver", "vize", "vor", "weg", "wett", "wider", "zer", "zu", "zurecht",
    "zurück", "zusammen", "zuwider", "zwischen"
]

circumfixes = [
    "be...ig", "be...t", "ge...e", "ge...ig", "ge...sel", "ge...t", "ver...ig"
]

suffixes = [
    "a", "abel", "ibel", "ade", "iade", "age", "aholic", "oholic", "oholiker", "aille", "al", "ell", "ament",
    "ement", "an", "and", "ant", "ent", "ante", "ente", "anz", "enz", "ar", "är", "arium", "arm", "artig", "ast", "at",
    "ee", "ei", "eierei", "el", "elchen", "erchen", "elle", "ens", "er", "erich", "erie", "ern", "esk", "ess", "esse",
    "isse", "ette", "eur", "euse", "fach", "fähig", "bold", "chen", "dings", "drom", "e", "i", "ian", "jan", "ice",
    "icht", "ie", "ier", "ieren", "ifizier", "isier", "iere", "ig","iges","iger","ige", "ik", "iker", "ine", "ing", "ingen", "en", "ern",
    "er", "e", "ell", "ion", "tion", "ation", "ismus", "asmus", "ist", "it", "ität", "itis", "iv", "ativ", "ke", "lei",
    "lein", "lekt", "ler", "lich", "ling", "lings", "lon", "los", "mals", "maßen", "mäßig", "mini", "n", "nis", "o",
    "oid", "ol", "or", "ator", "itor", "os", "ös", "ose", "ow", "pflichtig", "reich", "rich", "sal", "sam", "schaft",
    "sche", "seitig", "sel", "sen", "skop", "tex", "thek", "trächtig", "tum", "ung", "ur", "voll", "wang", "wangen",
    "wart", "wärts", "weg", "weise", "werk", "wesen", "zid","ete"
]

diminutive_suffixes = [
     'chen', 'lein', 'erl', 'el', 'rl', 'elein', 'i'
    ]
diminutive_suffix_exceptions={
    'chen':["pratze","pfotschen","Mannschen"],
    'lein':[],
    'erl':[],
    'el':["acht","butz"],
    'rl':[],
    'ele':["stiel"],
    'elein':[],
    'i':["sau","Martini","Matthäi","Johanni"]
}
komposita_exceptions=["Amizone","Sechsämterer","Bösartiges","Großartiger","Zungenbader","Eisenbahner","Eisenbahnerer","Eisenbähner","Baldes","Bempes","Haselnussblitzer"]
ableitung_exceptions=["Fremdarbeiter"]
"""
check to see if there need to be specific checks for specific suffixes:
elchen:removed, covered by chen 
chen: done
el: done
lein:done
erl:done
rl: done
ele: done
elein:done
i: done
Deverbale Kollektiva
pfotschen
letzte Lemma Erdäpfelquirl falsch zugeordnet
Paper aus dem Script machen ? (mit Manuel darüber reden)
"""


verb_diminutive_suffixes =[

]

umlautungen_und_ablautungen = {
    'a': ['ä', 'i', 'e', 'u', 'ie'],
    'e': ['a', 'o', 'ie', 'u'],
    'i': ['a', 'o', 'ie', 'u'],
    'o': ['e', 'ö', 'a', 'ie'],
    'u': ['o', 'ü', 'ie'],
    'au': ['äu', 'o', 'ie'],
    'ei': ['i', 'ie'],
    'ie': ['o'],
    'ö': ['o']
}

doppelkonsonanten = {
    'b': ['bb'],
    'd': ['dd'],
    'g': ['gg'],
    'k': ['kk'],
    'l': ['ll'],
    'm': ['mm'],
    'n': ['nn'],
    'p': ['pp'],
    'r': ['rr'],
    's': ['ss'],
    't': ['tt']
}


def read_data(name:str = 'tbl_eintraege.tab')->pd.DataFrame:
    """
    Diese Methode nimmt eine Eingabe und liest eine Datei mit diesem Namen aus dem gleichen Verzeichnis und gibt einen darauf basierenden Pandas Dataframe zurück
    """
    # Pfad zur CSV-Datei
    #hier relativen Pfad statt absoluten benutzen, dann muss lediglich die Datenbank im gleichen Ordner liegen
    csv_file = Path(__file__).parent.joinpath(name)

    rows = []

    with open(csv_file, encoding='utf-8') as file:
        for line in file:
            rows.append(line.strip().split('\t')[0:6])
    # DataFrame erstellen
    df = pd.DataFrame(rows[1:], columns=rows[0])
    return df


def find_ableitungen(data:pd.DataFrame)->dict:  
    """
    Funktionsweise:
        Erhalte einen pandas Dataframe, welcher mindestens die Spalten Grammatik, Grundform und Lemma enthalten muss. Finde alle Ableitungen der Grundformen des Dataframe und ordne diese in einem dict dem letzten Lemma zu.
        Dieses dict wird zurückgegeben. Es werden Explizit keine Diminuitiva erkannt.
    Eingabe:
        data:pd.DataFrame = dies ist ein pandas Dataframe, welcher die Daten enthalten soll. Diese Daten müssen mindestens Grundform, Lemma und Grammatik beinhalten.
    Ausgabe:
        Ableitungen:dict = dieses dictionary enthält alle Ableitungen des Dataframe data und ordnet diese dem letzten Lemma der Grundform zu.
    """
    Ableitungen=dict()
    data=data[['Grundform','Lemma']].drop_duplicates().groupby('Grundform')    #Dies generiert einen Dataframe, welcher zu jeder unterschiedlichen Grundform, alle dazugehörigen Lemma anzeigt
    for grundform, group in data:
        lemmata=group['Lemma'].tolist()
        ableitung,lemma = is_ableitung(grundform,lemmata) #Hier findet die Prüfung statt, ob es wirklich eine Ableitung ist, oder lediglich eine Grundform
        if ableitung:   
            if lemma in Ableitungen:
                Ableitungen[lemma].append(grundform)
            else:
                Ableitungen[lemma]=[grundform]
    return Ableitungen


def is_ableitung(grundform:str,lemmata:list)->tuple:
    """
    Funktionsweise:
        hier wird überprüft ob ein Wort mit gegebenen Lemmata eine Ableitung ist und gibt das letzte Lemma des Wortes zurück. Diminuitiva werden hier explizit ausgelassen.
    Eingabe:
        grundform:str= die Grundform des Wortes
        lemmata:list= Eine Liste mit allen Lemmata des Wortes
    Ausgabe:
        is_ableitung:boolean = Gibt an ob das Wort eine Ableitung ist
        last_lemma:str = das letzte Lemma des Wortes
    """
    #TODO: Komposita nicht mehr erkennen sondern nur noch Ableitugen!
    first_lemma,last_lemma = find_first_last_lemma(grundform,lemmata)
    if last_lemma=="":
        return False,""
    grundform = grundform.lower()
    lemmata = [lemma.lower() for lemma in lemmata]
    first_lemma=first_lemma.lower()
    last_lemma=last_lemma.lower()
    #lemmata = [lemma for lemma in lemmata if lemma not in prefixes]
    is_ableitung=False
    for prefix in prefixes:
        if grundform.startswith(prefix) and not first_lemma.startswith(prefix):
                is_ableitung = True
                break
    # Prüfen auf Zirkumfixe
    if not is_ableitung:
        for circumfix in circumfixes:
            parts = circumfix.split("...")
            if grundform.startswith(parts[0]) and grundform.endswith(parts[1]) and not last_lemma.endswith(parts[1]):
                if all(not ((x in grundform[len(parts[0]):]) ^ (x in grundform)) for x in lemmata):
                    is_ableitung = True
                    break
    # Prüfen auf Suffixe
    if not is_ableitung:
        for suffix in suffixes:
            if grundform.endswith(suffix) and (not last_lemma.endswith(suffix) or (last_lemma.endswith(suffix) and grundform[:-len(suffix)].endswith(suffix))):
                is_ableitung = True
                break
                
    if dim_checker(grundform,lemmata)[0]:
        return False,last_lemma
    
    if grundform in map(str.lower,ableitung_exceptions):
        is_ableitung=False
    
    return is_ableitung,last_lemma


def find_komposita(data:pd.DataFrame)->dict:
    """
    Funktionsweise:
        Erhalte einen pandas Dataframe, welcher mindestens die Spalten Grammatik, Grundform und Lemma enthalten muss. Finde alle Komposita der Grundformen des Dataframe und ordne diese in einem dict dem letzten Lemma zu.
        Dieses dict wird zurückgegeben
    Eingabe:
        data:pd.DataFrame = dies ist ein pandas Dataframe, welcher die Daten enthalten soll. Diese Daten müssen mindestens Grundform, Lemma und Grammatik beinhalten.
    Ausgabe:
        result:dict = dieses dictionary enthält alle Komposita des Dataframe data und ordnet diese dem letzten Lemma der Grundform zu.    
    """
    result=dict()
    data=data[["Grammatik","Lemma","Grundform"]]
    data=data.drop_duplicates()
    data=data[data["Grammatik"].map(lambda x: str(x).startswith('S'))]  #Entferne alle Zeilen, die keine Substantive sind, Namen werden hier implizit entfernt    
    data=data[data['Lemma'].map(lambda x:  (str(x) not in prefixes))] #Entferne alle Lemmata welche auch als prefixe fungieren (Hierbei nur die Prefixe in der Prefix Liste, diese sollte bei Bedarf up to date gehalten werden)
    data=data[['Grundform','Lemma']].groupby('Grundform').filter(lambda x: len(x)>1).drop_duplicates()
    for grundform,group in data.groupby("Grundform"):
        for lemmata in homographs(grundform,group['Lemma'].tolist()):
            if len(lemmata)>1:
                lemma=find_last_lemma(grundform,lemmata)
                if " " not in grundform and\
                "-" not in grundform and\
                len(lemmata)>1 and\
                grundform not in komposita_exceptions and\
                    not is_ableitung(grundform, lemmata)[0] and\
                        not dim_checker(grundform,lemmata)[0]: 
                    if lemma in result:
                        result[lemma].append(grundform)
                    else:
                        result[lemma]=[grundform]
    return result

    


def find_hapax_legomena(data:pd.DataFrame)->list[str]:
    """
    Funktionsweise:
        Findet alle Hapax Legomena eines Dataframe und gibt diese zurück
    Eingabe:
        data:pd.DataFrame= ein Dataframe, welcher mindestens Grundform und Lemma enthalten muss
    Ausgabe:
        result:list[str] = eine Liste mit allen Hapax Legomena des Dataframe
    """
    result=data[['Grundform','Lemma']].groupby(['Grundform','Lemma']).filter(lambda x: len(x) == 1)['Grundform'].drop_duplicates().tolist()
    return result


def find_diminuitive(data:pd.DataFrame)->dict:
    """
    Funktionsweise:
        Erhalte einen pandas Dataframe, welcher mindestens die Spalten Grammatik, Grundform und Lemma enthalten muss. Finde alle Diminuitiva der Grundformen des Dataframe und ordne diese in einem dict dem letzten Lemma zu.
        Dieses dict wird zurückgegeben
    Eingabe:
        data:pd.DataFrame = dies ist ein pandas Dataframe, welcher die Daten enthalten soll. Diese Daten müssen mindestens Grundform, Lemma und Grammatik beinhalten.
    Ausgabe:
        Diminuitiva:dict = dieses dictionary enthält alle Diminuitiva des Dataframe data und ordnet diese dem letzten Lemma der Grundform zu.
    """
    Diminuitiva=dict()
    data=data[data["Grammatik"].map(lambda x: str(x).startswith('S'))]  #Entferne alle Zeilen, die keine Substantive sind, Namen werden hier implizit entfernt
    data=data[['Grundform','Lemma']].drop_duplicates().groupby('Grundform')    #Dies generiert einen Dataframe, welcher zu jeder unterschiedlichen Grundform, alle Dazugehörigen Lemma anzeigt
    for grundform, group in data:
        lemmata=group['Lemma'].tolist()
        if len(lemmata)<1:
            continue
        diminuitiv,lemma,end=dim_checker(grundform,lemmata)
        if diminuitiv:
            if end in Diminuitiva:
                if lemma in Diminuitiva[end]:
                    Diminuitiva[end][lemma].append(grundform)
                else:
                    Diminuitiva[end][lemma]=[grundform]
            else:
                Diminuitiva[end]={lemma: [grundform] }
    return Diminuitiva

def find_last_lemma(grundform:str,lemmata:list)->str:
    """
    Funktionsweise:
        Finde das letzte Lemma eines Wortes, bei Angabe der Lemmata des Wortes.
    Eingabe:
        grundform:str = dies ist die grundform des Wortes
        lemmata:list = dies ist die Liste aller Lemmata des Wortes 
    Ausgabe:
        last_lemma:str = dies ist der String des gefundenen letzten Lemma, wenn das Wort keine Lemma hat, dann ist dieser String leer. 
    """
    grundform=grundform.lower()
 
    #check which lemma comes last in the word
    index=-1
    wort=grundform
    if len(lemmata)<1:
        print("there was something wrong with the lemmata")
        print(grundform)
        print(lemmata)
        return ""
    last_lemma=lemmata[0]
    if len(lemmata)>1:
        for Lemma in lemmata:
            lemma=Lemma.lower()
            if lemma in wort:
                ind=wort.rfind(lemma)
                if ind+len(lemma) > index:
                    index=ind
                    last_lemma=Lemma
            else:
                temp=lemma
                while len(temp)>1:
                    temp=temp[:-1]
                    if temp in wort:
                        ind=wort.rfind(temp)
                        if ind+len(temp) > index:
                            index=ind
                            last_lemma=Lemma
                        break
    #TODO: Umlaute am Lemmaanfang werden nicht erkannt    
    return last_lemma

def find_first_last_lemma(grundform:str,lemmata:list)->tuple:
    """
    Funktionsweise:
        Finde das letzte und erste Lemma eines Wortes, bei Angabe der Lemmata des Wortes.
    Eingabe:
        grundform:str = dies ist die grundform des Wortes
        lemmata:list = dies ist die Liste aller Lemmata des Wortes 
    Ausgabe:
        last_lemma:str = dies ist der String des gefundenen letzten Lemma, wenn das Wort keine Lemma hat, dann ist dieser String leer. 
        first_lemma:str = dies ist der String des gefundenen ersten Lemma, wenn das Wort keine Lemma hat, dann ist dieser String leer.
    """
    grundform=grundform.lower()
 
    #check which lemma comes last in the word
    lindex=len(grundform)
    rindex=-1
    wort=grundform
    if len(lemmata)<1:
        print("there was something wrong with the lemmata")
        print(grundform)
        print(lemmata)
        return ""
    last_lemma=lemmata[0]
    first_lemma=lemmata[0]
    if len(lemmata)>1:
        for Lemma in lemmata:
            lemma=Lemma.lower()
            if lemma in wort:
                rind=wort.rfind(lemma)
                lind=wort.find(lemma)
                if rind+len(lemma) > rindex:
                    rindex=rind
                    last_lemma=Lemma
                if lind < lindex:
                    lindex=lind
                    first_lemma=Lemma
            else:
                temp=lemma
                while len(temp)>1:
                    temp=temp[:-1]
                    if temp in wort:
                        rind=wort.rfind(temp)
                        lind=wort.find(temp)
                        if rind+len(temp) > rindex:
                            rindex=rind
                            last_lemma=Lemma
                        if lind < lindex:
                            lindex=lind
                            first_lemma=Lemma
                        break
                    #Das könnte zu falschen Erkennungen führen, muss unter Beobachtung bleiben.
    #TODO: Umlaute am Lemmaanfang werden nicht erkannt   
    return (first_lemma,last_lemma)

def dim_checker(grundform:str,lemmata:list)->tuple: 
    """
    Funktionsweise:
        hier wird überprüft ob ein Wort mit gegebenen Lemmata ein Diminuitiv ist und gibt das letzte Lemma des Wortes zurück.
    Eingabe:
        grundform:str= die Grundform des Wortes
        lemmata:list= Eine Liste mit allen Lemmata des Wortes
    Ausgabe:
        diminuitiv:boolean = Gibt an ob das Wort ein Diminuitiv ist
        last_lemma:str = das letzte Lemma des Wortes
    
    """
    grundform=grundform.lower()
 
    #check which lemma comes last in the word
    last_lemma=find_last_lemma(grundform,lemmata)
    #das letzte Lemma wurde gefunden
    if len(last_lemma)<1:
        return False,[],""
    end=""
    diminuitiv=False
    if " " not in grundform and "-" not in grundform: #hier werden grundformen aus mehreren Worten oder Worten mit Bindestrichen entfernt
        for suffix in diminutive_suffixes:
            if grundform in diminutive_suffix_exceptions[suffix] or last_lemma in diminutive_suffix_exceptions[suffix]:
                continue
            if grundform.endswith(suffix) and not last_lemma.endswith(suffix): #check ob das Suffix echt ist und nicht Teil des letzten Lemma
                end=suffix
                diminuitiv=True
                if suffix=="chen":
                    amount=0
                    for lemma in lemmata:
                        amount += lemma.lower().count("ch")
                    if grundform.count("ch") <= amount:
                        end=""
                        diminuitiv=False                        
                if suffix == "el":
                    if last_lemma.endswith("eln") or (last_lemma[0].isupper() and last_lemma.endswith("en")):
                        end=""
                        diminuitiv=False
                if suffix == "ele" and last_lemma.endswith("elen"):
                    end=""
                    diminuitiv=False
                if suffix == "i" and (last_lemma.endswith("en") or last_lemma.endswith("ei")or last_lemma.endswith("ai")):
                    end=""
                    diminuitiv=False
                if diminuitiv:
                    break
    
    return diminuitiv,last_lemma,end

def homographs(grundform:str,lemmata:list)->list[list]:
    homograph=[[]]
    length = sum(len(x) for x in lemmata)
    if length >= len(grundform)*2: #Das Wort ist ein Homograph TODO
        if len(lemmata)==2:
            homograph=[[lemmata[0]]]
            homograph.append([lemmata[1]])
        else:
            for lemma in lemmata:
                #TODO!!!! Das funktioniert nicht bei Worten mit mehreren Lemmata
                if homograph==[[]]:
                    homograph=[[lemma]]
                else:
                    homograph.append([lemma])
        #print(grundform)
    else:
        homograph=[lemmata]
    return homograph

def ui_ableitungen():
    save = FD.asksaveasfilename(filetypes=[('excel files','*.xlsx *.xlsm *xlsb *.xltx *.xls *.xlt *.xml *.xlam *.xla *.xlw *.xlr')],defaultextension='xlsx')
    data=df
    ableitungen=find_ableitungen(data)

    max_zeilen = max(len(liste) for liste in ableitungen.values())
    # Excel-Datei erstellen
    wb = Workbook()
    ws = wb.active
    ws.title = "Ableitungen"
    ableitungen=dict(sorted(ableitungen.items()))
    # Spaltenüberschriften schreiben
    headers = list(ableitungen.keys())
    ws.append(headers)

    # Zeilen schreiben, fehlende Werte mit None auffüllen
    for i in range(max_zeilen):
        zeile = [ableitungen[spalte][i] if i < len(ableitungen[spalte]) else None for spalte in headers]
        ws.append(zeile)

    # Datei speichern
    wb.save(save)
    tkinter.messagebox.showinfo("Erflogreich durchgeführt",  "Ableitungen erfolgreich gespeichert!")



def ui_hapax_legomena():
    save = FD.asksaveasfilename(filetypes=[('excel files','*.xlsx *.xlsm *xlsb *.xltx *.xls *.xlt *.xml *.xlam *.xla *.xlw *.xlr')],defaultextension='xlsx')
    data=df
    hapax=find_hapax_legomena(data)
    hapax.sort()
    # Excel-Datei erstellen
    wb = Workbook()
    ws = wb.active
    ws.title = "Hapax Legomena"

    for item in hapax:
        ws.append([item])

    # Zeilen schreiben, fehlende Werte mit None auffüllen
    wb.save(save)
    tkinter.messagebox.showinfo("Erflogreich durchgeführt",  "Hapax Legomena erfolgreich gespeichert!")

def ui_Diminuitiva():
    save = FD.asksaveasfilename(filetypes=[('excel files','*.xlsx *.xlsm *xlsb *.xltx *.xls *.xlt *.xml *.xlam *.xla *.xlw *.xlr')],defaultextension='xlsx')
    data=df
    All_diminuitiva=find_diminuitive(data)
    wb=Workbook()
    for diminuitiva in All_diminuitiva:
        max_zeilen = max(len(liste) for liste in All_diminuitiva[diminuitiva].values())
        ws = wb.create_sheet(diminuitiva)
        dims=All_diminuitiva[diminuitiva]
        dims=dict(sorted(dims.items()))
        headers = list(dims.keys())
        ws.append(headers)
        for i in range(max_zeilen):
            zeile = [dims[spalte][i] if i < len(dims[spalte]) else None for spalte in headers]
            ws.append(zeile)
    wb.remove(wb[wb.sheetnames[0]])
    wb.save(save)
    tkinter.messagebox.showinfo("Erflogreich durchgeführt",  "Diminuitiva erfolgreich gespeichert!")

def ui_Komposita():
    save = FD.asksaveasfilename(filetypes=[('excel files','*.xlsx *.xlsm *xlsb *.xltx *.xls *.xlt *.xml *.xlam *.xla *.xlw *.xlr')],defaultextension='xlsx')
    data=df
    Komposita = find_komposita(data)
    Komposita = dict(sorted(Komposita.items()))
    max_zeilen = max(len(liste) for liste in Komposita.values())
    wb= Workbook()
    ws = wb.active
    ws.title="Komposita"

    headers= list(Komposita.keys())
    ws.append(headers)

    for i in range(max_zeilen):
        zeile = [Komposita[spalte][i] if i < len(Komposita[spalte]) else None for spalte in headers]
        ws.append(zeile)
    wb.save(save)
    tkinter.messagebox.showinfo("Erflogreich durchgeführt",  "Komposita erfolgreich gespeichert!")

def ui_last_lemma():
    save = FD.asksaveasfilename(filetypes=[('excel files','*.xlsx *.xlsm *xlsb *.xltx *.xls *.xlt *.xml *.xlam *.xla *.xlw *.xlr')],defaultextension='xlsx')
    data=df
    letzte_lemma={}
    data=data[['Grundform','Lemma']].drop_duplicates().groupby('Grundform')
    for grundform,group in data:
        lemmata=group['Lemma'].tolist()
        lemma= find_last_lemma(grundform,lemmata)
        if lemma in letzte_lemma:
            letzte_lemma[lemma].append(grundform)
        else:
            letzte_lemma[lemma]=[grundform]
    max_zeilen = max(len(liste) for liste in letzte_lemma.values())
    wb=Workbook()
    ws = wb.active
    ws.title="Letzte Lemma"
    letzte_lemma=dict(sorted(letzte_lemma.items()))
    headers = list(letzte_lemma.keys())
    ws.append(headers)
    for i in range(max_zeilen):
        zeile = [letzte_lemma[spalte][i] if i < len(letzte_lemma[spalte])else None for spalte in headers]
        ws.append(zeile)
    wb.save(save)
    tkinter.messagebox.showinfo("Erflogreich durchgeführt",  "Letzte Lemma erfolgreich gespeichert!")

    

def ui_readinput():
    csv_file=FD.askopenfilename()
    rows = []
    
    with open(csv_file,encoding='utf-8') as file:
        for line in file:
            rows.append(line.strip().split('\t')[0:6])
    # DataFrame erstellen
    global df
    df = pd.DataFrame(rows[1:], columns=rows[0])
    tkinter.messagebox.showinfo("Erflogreich durchgeführt",  "Eingabedatenbank erfolgreich eingelesen!")


def test1(name="Ableitungen.csv"):
    tmp_time=time.time()
    data= read_data()
    print("Finde Ableitungen")
    print("-----------------")
    ableitungen=find_ableitungen(data)
    print("Ableitungen gefunden")
    print("-----------------")
    print("Ableitungen werden in "+name+" gespeichert")
    print("-----------------")
    max_zeilen = max(len(liste) for liste in ableitungen.values())

    # Excel-Datei erstellen
    wb = Workbook()
    ws = wb.active
    ws.title = "Ableitungen"

    # Spaltenüberschriften schreiben
    headers = list(ableitungen.keys())
    ws.append(headers)

    # Zeilen schreiben, fehlende Werte mit None auffüllen
    for i in range(max_zeilen):
        zeile = [ableitungen[spalte][i] if i < len(ableitungen[spalte]) else None for spalte in headers]
        ws.append(zeile)

    # Datei speichern
    wb.save("Ableitungen.xlsx")
    with open(name,mode='w',encoding="utf-8" ,errors='replace' ) as file:
        writer=csv.writer(file)
        writer.writerow(["Lemma","Grundform"])
        for key,value in ableitungen.items():
            writer.writerow([key,value])
    print("Ableitungen gespeichert")
    print("-----------------")
    print("Ausführungsdauer Test 1: %s Sekunden" %(time.time()-tmp_time))
    print("-----------------")    

def test2(name="hapax_legomena.csv"):
    tmp_time=time.time()
    data=read_data()
    print("Finde Hapax Legomena")
    print("-----------------")
    hapax=find_hapax_legomena(data)
    print("Hapax Legomena gefunden")
    print("-----------------")
    print("Hapax Legomena werden in " +name+" gespeichert" )
    print("-----------------")

    with open(name, mode='w',encoding='utf-8',errors='replace') as file:
        writer=csv.writer(file)
        for value in hapax:
            writer.writerow([value])
    print("Hapax Legomena geschrieben")
    print("-----------------")
    print("Ausführungsdauer Test 2: %s Sekunden" %(time.time()-tmp_time))
    print("-----------------")    

def test3(name="Diminuitiva.xlsx"):
    tmp_time=time.time()
    data=read_data()
    print("Finde Diminuitiva")
    print("-----------------")
    All_diminuitiva=find_diminuitive(data)
    print("Diminuitiva gefunden")
    print("-----------------")
    print("Diminuitiva werden in " +name+" gespeichert" )
    print("-----------------")
    wb=Workbook()
    for diminuitiva in All_diminuitiva:
        max_zeilen = max(len(liste) for liste in All_diminuitiva[diminuitiva].values())
        # Excel-Datei erstellen
        ws = wb.create_sheet(diminuitiva)

        # Spaltenüberschriften schreiben
        headers = list(All_diminuitiva[diminuitiva].keys())
        ws.append(headers)

        # Zeilen schreiben, fehlende Werte mit None auffüllen
        for i in range(max_zeilen):
            zeile = [All_diminuitiva[diminuitiva][spalte][i] if i < len(All_diminuitiva[diminuitiva][spalte]) else None for spalte in headers]
            ws.append(zeile)

        # Datei speichern
    wb.remove(wb[wb.sheetnames[0]])
    wb.save(name)
   
    """
    with open(name, mode='w',encoding='utf-8',errors='replace') as file:
        writer=csv.writer(file)
        writer.writerow(["Lemma","Grundform"])
        for key,value in diminuitiva.items():
            writer.writerow([key,value])"""
    print("Diminuitiva geschrieben")
    print("-----------------")
    print("Ausführungsdauer Test 3: %s Sekunden" %(time.time()-tmp_time))
    print("-----------------")    

def test4(name="Komposita.csv"):
    tmp_time=time.time()
    data=read_data()
    print("Finde Komposita")
    print("-----------------")
    Komposita=find_komposita(data)
    print("Komposita gefunden")
    print("-----------------")
    print("Komposita werden in " +name+" gespeichert" )
    print("-----------------")
    max_zeilen = max(len(liste) for liste in Komposita.values())
    # Excel-Datei erstellen
    wb = Workbook()
    ws = wb.active
    ws.title = "Komposita"

    # Spaltenüberschriften schreiben
    headers = list(Komposita.keys())
    ws.append(headers)

    # Zeilen schreiben, fehlende Werte mit None auffüllen
    for i in range(max_zeilen):
        zeile = [Komposita[spalte][i] if i < len(Komposita[spalte]) else None for spalte in headers]
        ws.append(zeile)

    # Datei speichern
    wb.save("Komposita.xlsx")
    with open(name, mode='w',encoding='utf-8',errors='replace') as file:
        writer=csv.writer(file)
        writer.writerow(["Lemma","Grundform"])
        for key,value in Komposita.items():
            writer.writerow([key,value])
    print("Komposita geschrieben")
    print("-----------------")
    print("Ausführungsdauer Test 4: %s Sekunden" %(time.time()-tmp_time))
    print("-----------------")    

def test5(name="last_lemma.csv"):
    tmp_time=time.time()
    data=read_data()
    print("Finde letzte Lemma")
    print("-----------------")
    letzte_lemma={}
    data=data[['Grundform','Lemma']].drop_duplicates().groupby('Grundform')
    for grundform,group in data:
        lemmata=group['Lemma'].tolist()
        lemma=find_last_lemma(grundform,lemmata)
        if lemma in letzte_lemma:
            letzte_lemma[lemma].append(grundform)
        else:
            letzte_lemma[lemma]=[grundform]
    print("Letzte Lemma zugeordnet")
    print("-----------------")
    print("Letzte Lemma werden in " +name+" gespeichert" )

    print("-----------------")
    max_zeilen = max(len(liste) for liste in letzte_lemma.values())
    wb = Workbook()
    ws = wb.active
    ws.title = "Letzte_Lemma"

    # Spaltenüberschriften schreiben
    headers = list(letzte_lemma.keys())
    ws.append(headers)

    # Zeilen schreiben, fehlende Werte mit None auffüllen
    for i in range(max_zeilen):
        zeile = [letzte_lemma[spalte][i] if i < len(letzte_lemma[spalte]) else None for spalte in headers]
        ws.append(zeile)

    wb.save("letzte_lemma.xlsx")


    with open(name, mode='w',encoding='utf-8',errors='replace') as file:
        writer=csv.writer(file)
        for value in letzte_lemma:
            writer.writerow([value])

def user_interface(): 
    eingabefeld = Button(fenster,text="Eingabedatenbank",command=ui_readinput)

    ableitungen_button = Button(fenster,text="Ableitungen",command = ui_ableitungen)
    hapax_button = Button(fenster, text="Hapax Legomena", command = ui_hapax_legomena)
    diminuitiva_button = Button(fenster, text="Diminuitiva", command = ui_Diminuitiva)
    komposita_button = Button(fenster, text = "Komposita", command = ui_Komposita)
    last_lemma_button = Button(fenster, text="Letzte Lemma", command = ui_last_lemma)

    exit_button = Button(fenster, text="Beenden",command=fenster.quit)

    eingabefeld.pack(side=TOP)
    ableitungen_button.pack(side=LEFT)
    hapax_button.pack(side=LEFT)
    diminuitiva_button.pack(side=LEFT)
    komposita_button.pack(side=LEFT)
    last_lemma_button.pack(side=LEFT)
    exit_button.pack(side=BOTTOM)

    mainloop()





if __name__=="__main__":

    start_time = time.time()
    user_interface()
    #print("Execute")
    #test1()
    #test2()
    #test3()
    #test4()
    #test5()
    #print("Done")
    #print("Gesamtdauer: %s Sekunden" %(time.time()-start_time))

class User_Interface(tkinter.Frame):
    def __init__(self,master=None):
        super().___init__(master)
        self.pack()
        self.createWidgets()

    def createWidgets(self):
        self.nameEntry=tkinter.Entry(self)
        self.nameEntry.pack()




"""
Notizen:
    Ableitungen:

    Komposita:
        Grenzfall zusammenrückungen wie "Heimackern"
        Sonderfälle: "Amizone","Sechsämterer","Bösartiges","Großartiger","Zungenbader","Eisenbahner","Eisenbahnerer","Eisenbähner","Baldes","Bempes","Haselnussblitzer"
        Sechsämterer ist Ableitung
        Homographe werden fälschlich Erkannt. "Idee mit der Länge der Lemmata"
        Arbeiter keine Komposita,da Zusammenbildung mit Ausnahme des Fremdarbeiter
        -iges, -iger, -ige endung sind Ableitungen
        -ete in Besonderer Form Augete sind Ableitungen Beinete
        Wortbildungen aus bisherigen Komposita müssen Ableitungen sein <- sollte implizit gelößt sein durch das exkludieren von Ableitungen
        -ler ableitungen, aus 


Worte in den Ableitungen sind auch in den Komposita, sollte nicht sein

"""